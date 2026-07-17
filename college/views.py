from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Sum, Q
from django.http import HttpResponse, Http404
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from .models import UserProfile, Student, Attendance, FeeTransaction, AccountTransaction, SalaryPayment
from .forms import StudentForm, LeaveForm, FeePaymentForm, AccountTransactionForm, SalaryPaymentForm, StaffCreateForm, StaffEditForm, StaffPasswordForm

# Custom Decorator for Admin (Super Admin)
def is_admin(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return hasattr(user, 'profile') and user.profile.role == 'ADMIN'

def is_staff(user):
    return user.is_authenticated and hasattr(user, 'profile')

# ----------------- AUTH VIEWS -----------------
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f"Welcome back, {user.first_name or user.username}!")
                return redirect('dashboard')
        else:
            messages.error(request, "Invalid username or password.")
    else:
        form = AuthenticationForm()
    return render(request, 'college/login.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.info(request, "Logged out successfully.")
    return redirect('login')


# ----------------- STUDENT MANAGEMENT -----------------
@user_passes_test(is_admin, login_url='dashboard')
def student_list(request):
    class_filter = request.GET.get('class_name', '')
    batch_filter  = request.GET.get('batch', '')
    search_query  = request.GET.get('search', '')

    students = Student.objects.filter(is_active=True).order_by('class_name', 'roll_number')
    if class_filter:
        students = students.filter(class_name__iexact=class_filter)
    if batch_filter:
        students = students.filter(batch__iexact=batch_filter)
    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)  |
            Q(roll_number__icontains=search_query) |
            Q(admission_number__icontains=search_query)
        )

    classes = Student.objects.values_list('class_name', flat=True).distinct()
    batches = Student.objects.values_list('batch', flat=True).distinct()

    context = {
        'students': students,
        'classes': classes,
        'batches': batches,
        'class_filter': class_filter,
        'batch_filter': batch_filter,
        'search_query': search_query,
        'total_count': students.count(),
    }
    return render(request, 'college/student_list.html', context)


@user_passes_test(is_admin, login_url='dashboard')
def student_add(request):
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            student = form.save()
            messages.success(request, f"Student '{student.full_name}' added successfully!")
            return redirect('student_list')
    else:
        form = StudentForm()
    return render(request, 'college/student_form.html', {
        'form': form,
        'action': 'Add New',
        'btn_label': 'Add Student',
    })


@user_passes_test(is_admin, login_url='dashboard')
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, f"Student '{student.full_name}' updated successfully!")
            return redirect('student_list')
    else:
        form = StudentForm(instance=student)
    return render(request, 'college/student_form.html', {
        'form': form,
        'student': student,
        'action': 'Edit',
        'btn_label': 'Save Changes',
    })


@user_passes_test(is_admin, login_url='dashboard')
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        name = student.full_name
        student.delete()
        messages.success(request, f"Student '{name}' has been deleted.")
        return redirect('student_list')
    return render(request, 'college/student_confirm_delete.html', {'student': student})


# ----------------- STAFF MANAGEMENT (Admin only) -----------------
@user_passes_test(is_admin, login_url='dashboard')
def staff_list(request):
    staff_members = User.objects.filter(profile__role='STAFF').select_related('profile').order_by('first_name')
    return render(request, 'college/staff_list.html', {'staff_members': staff_members})


@user_passes_test(is_admin, login_url='dashboard')
def staff_add(request):
    if request.method == 'POST':
        form = StaffCreateForm(request.POST)
        if form.is_valid():
            d = form.cleaned_data
            user = User.objects.create_user(
                username=d['username'],
                password=d['password'],
                first_name=d['first_name'],
                last_name=d['last_name'],
                email=d.get('email', ''),
            )
            UserProfile.objects.create(
                user=user,
                role='STAFF',
                phone=d.get('phone', ''),
                salary=d['salary'],
            )
            messages.success(request, f"Staff account for '{user.get_full_name()}' created successfully!")
            return redirect('staff_list')
    else:
        form = StaffCreateForm()
    return render(request, 'college/staff_form.html', {
        'form': form,
        'action': 'Add New',
        'btn_label': 'Create Staff Account',
    })


@user_passes_test(is_admin, login_url='dashboard')
def staff_edit(request, pk):
    staff_user = get_object_or_404(User, pk=pk, profile__role='STAFF')
    profile = staff_user.profile
    if request.method == 'POST':
        form = StaffEditForm(request.POST)
        if form.is_valid():
            d = form.cleaned_data
            staff_user.first_name = d['first_name']
            staff_user.last_name  = d['last_name']
            staff_user.email      = d.get('email', '')
            staff_user.save()
            profile.phone  = d.get('phone', '')
            profile.salary = d['salary']
            profile.save()
            messages.success(request, f"'{staff_user.get_full_name()}' profile updated successfully.")
            return redirect('staff_list')
    else:
        form = StaffEditForm(initial={
            'first_name': staff_user.first_name,
            'last_name':  staff_user.last_name,
            'email':      staff_user.email,
            'phone':      profile.phone,
            'salary':     profile.salary,
        })
    return render(request, 'college/staff_form.html', {
        'form': form,
        'staff_user': staff_user,
        'action': 'Edit',
        'btn_label': 'Save Changes',
    })


@user_passes_test(is_admin, login_url='dashboard')
def staff_set_password(request, pk):
    staff_user = get_object_or_404(User, pk=pk, profile__role='STAFF')
    if request.method == 'POST':
        form = StaffPasswordForm(request.POST)
        if form.is_valid():
            staff_user.set_password(form.cleaned_data['new_password'])
            staff_user.save()
            messages.success(request, f"Password for '{staff_user.get_full_name()}' changed successfully.")
            return redirect('staff_list')
    else:
        form = StaffPasswordForm()
    return render(request, 'college/staff_password.html', {
        'form': form,
        'staff_user': staff_user,
    })


@user_passes_test(is_admin, login_url='dashboard')
def staff_toggle_active(request, pk):
    """Activate / deactivate a staff login account."""
    staff_user = get_object_or_404(User, pk=pk, profile__role='STAFF')
    if request.method == 'POST':
        staff_user.is_active = not staff_user.is_active
        staff_user.save()
        state = 'activated' if staff_user.is_active else 'deactivated'
        messages.success(request, f"Account for '{staff_user.get_full_name()}' has been {state}.")
    return redirect('staff_list')


# ----------------- DASHBOARD -----------------
@login_required
def dashboard(request):
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else 'STAFF'
    
    today = timezone.localtime(timezone.now()).date()
    
    # Common stats
    total_students = Student.objects.filter(is_active=True).count()
    
    # Today's attendance summary
    total_marked_today = Attendance.objects.filter(date=today, period='Daily').count()
    present_today = Attendance.objects.filter(date=today, period='Daily', status='PRESENT').count()
    absent_today = Attendance.objects.filter(date=today, period='Daily', status='ABSENT').count()
    leave_today = Attendance.objects.filter(date=today, period='Daily', status='LEAVE').count()
    
    attendance_rate = 0
    if total_marked_today > 0:
        attendance_rate = round((present_today / total_marked_today) * 100, 1)
    
    # Financial Stats (Super Admin only)
    cash_in_hand = 0
    cash_in_bank = 0
    defaulters_count = 0
    income_today = 0
    expense_today = 0
    
    if role == 'ADMIN':
        # Calculate Cash in Hand & Cash in Bank
        income_cash = AccountTransaction.objects.filter(transaction_type='INCOME', payment_method='CASH').aggregate(total=Sum('amount'))['total'] or 0
        expense_cash = AccountTransaction.objects.filter(transaction_type='EXPENSE', payment_method='CASH').aggregate(total=Sum('amount'))['total'] or 0
        cash_in_hand = income_cash - expense_cash
        
        income_bank = AccountTransaction.objects.filter(transaction_type='INCOME', payment_method='BANK').aggregate(total=Sum('amount'))['total'] or 0
        expense_bank = AccountTransaction.objects.filter(transaction_type='EXPENSE', payment_method='BANK').aggregate(total=Sum('amount'))['total'] or 0
        cash_in_bank = income_bank - expense_bank
        
        # Defaulters Count: Students whose paid fees < total course fees
        # (Using a simpler python-based count or query filter if database-efficient)
        all_students = Student.objects.filter(is_active=True)
        defaulters_count = sum(1 for s in all_students if s.pending_fee > 0)
        
        # Incomes and Expenses today
        income_today = AccountTransaction.objects.filter(transaction_type='INCOME', date=today).aggregate(total=Sum('amount'))['total'] or 0
        expense_today = AccountTransaction.objects.filter(transaction_type='EXPENSE', date=today).aggregate(total=Sum('amount'))['total'] or 0
        
    # Build chart data for the last 6 months (Admin only)
    chart_months = []
    chart_income = []
    chart_expense = []
    
    if role == 'ADMIN':
        curr_date = today
        for i in range(5, -1, -1):
            target_month = curr_date - timedelta(days=30 * i)
            month_str = target_month.strftime("%B")
            month_num = target_month.month
            year_num = target_month.year
            
            inc = AccountTransaction.objects.filter(transaction_type='INCOME', date__month=month_num, date__year=year_num).aggregate(total=Sum('amount'))['total'] or 0
            exp = AccountTransaction.objects.filter(transaction_type='EXPENSE', date__month=month_num, date__year=year_num).aggregate(total=Sum('amount'))['total'] or 0
            
            chart_months.append(month_str[:3])
            chart_income.append(float(inc))
            chart_expense.append(float(exp))
            
    recent_transactions = AccountTransaction.objects.order_ok = AccountTransaction.objects.all().order_by('-date')[:5] if role == 'ADMIN' else []
    recent_leaves = Attendance.objects.filter(status='LEAVE').order_by('-date')[:5]
    
    context = {
        'role': role,
        'total_students': total_students,
        'attendance_rate': attendance_rate,
        'present_today': present_today,
        'absent_today': absent_today,
        'leave_today': leave_today,
        'cash_in_hand': cash_in_hand,
        'cash_in_bank': cash_in_bank,
        'total_balance': cash_in_hand + cash_in_bank,
        'defaulters_count': defaulters_count,
        'income_today': income_today,
        'expense_today': expense_today,
        'chart_months': chart_months,
        'chart_income': chart_income,
        'chart_expense': chart_expense,
        'recent_transactions': recent_transactions,
        'recent_leaves': recent_leaves,
    }
    return render(request, 'college/dashboard.html', context)


# ----------------- ATTENDANCE -----------------
@login_required
def attendance_list(request):
    today = timezone.localtime(timezone.now()).date()
    date_str = request.GET.get('date', today.strftime('%Y-%m-%d'))
    selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    
    class_filter = request.GET.get('class_name', '')
    batch_filter = request.GET.get('batch', '')
    period_filter = request.GET.get('period', 'Daily')
    
    students = Student.objects.filter(is_active=True)
    if class_filter:
        students = students.filter(class_name__iexact=class_filter)
    if batch_filter:
        students = students.filter(batch__iexact=batch_filter)
        
    classes = Student.objects.values_list('class_name', flat=True).distinct()
    batches = Student.objects.values_list('batch', flat=True).distinct()
    
    # Get existing attendances for this date and period
    existing_attendances = Attendance.objects.filter(date=selected_date, period=period_filter)
    attendance_map = {att.student_id: att.status for att in existing_attendances}
    reason_map = {att.student_id: att.leave_reason or '' for att in existing_attendances}
    
    if request.method == 'POST':
        for student in students:
            status_val = request.POST.get(f'status_{student.id}', 'PRESENT')
            leave_reason = request.POST.get(f'reason_{student.id}', '')
            
            Attendance.objects.update_or_create(
                student=student,
                date=selected_date,
                period=period_filter,
                defaults={
                    'status': status_val,
                    'leave_reason': leave_reason if status_val == 'LEAVE' else '',
                    'marked_by': request.user
                }
            )
        messages.success(request, f"Attendance saved successfully for {selected_date} ({period_filter}).")
        return redirect(f"{request.path}?date={date_str}&class_name={class_filter}&batch={batch_filter}&period={period_filter}")
        
    student_list = []
    for s in students:
        student_list.append({
            'id': s.id,
            'full_name': s.full_name,
            'roll_number': s.roll_number,
            'class_name': s.class_name,
            'batch': s.batch,
            'status': attendance_map.get(s.id, 'PRESENT'),
            'leave_reason': reason_map.get(s.id, ''),
        })
        
    context = {
        'students': student_list,
        'classes': classes,
        'batches': batches,
        'selected_date': selected_date,
        'date_str': date_str,
        'class_filter': class_filter,
        'batch_filter': batch_filter,
        'period_filter': period_filter,
        'periods': ['Daily', 'Period 1', 'Period 2', 'Period 3', 'Period 4', 'Period 5']
    }
    return render(request, 'college/attendance.html', context)

@login_required
def mark_leave(request):
    if request.method == 'POST':
        form = LeaveForm(request.POST)
        if form.is_valid():
            attendance = form.save(commit=False)
            attendance.marked_by = request.user
            # Check unique_together conflict
            conflict = Attendance.objects.filter(student=attendance.student, date=attendance.date, period=attendance.period).first()
            if conflict:
                conflict.status = 'LEAVE'
                conflict.leave_reason = attendance.leave_reason
                conflict.marked_by = request.user
                conflict.save()
            else:
                attendance.save()
            messages.success(request, f"Leave logged successfully for {attendance.student.full_name} on {attendance.date}.")
            return redirect('attendance_list')
    else:
        initial_student = request.GET.get('student_id', None)
        form = LeaveForm(initial={'date': timezone.now().date(), 'student': initial_student})
    return render(request, 'college/mark_leave.html', {'form': form})


# ----------------- FEES COLLECTION -----------------
@user_passes_test(is_admin, login_url='dashboard')
def fee_collection(request):
    class_filter = request.GET.get('class_name', '')
    batch_filter = request.GET.get('batch', '')
    search_query = request.GET.get('search', '')
    
    students = Student.objects.filter(is_active=True)
    if class_filter:
        students = students.filter(class_name__iexact=class_filter)
    if batch_filter:
        students = students.filter(batch__iexact=batch_filter)
    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(roll_number__icontains=search_query) |
            Q(admission_number__icontains=search_query)
        )
        
    classes = Student.objects.values_list('class_name', flat=True).distinct()
    batches = Student.objects.values_list('batch', flat=True).distinct()
    
    form = FeePaymentForm()
    
    context = {
        'students': students,
        'classes': classes,
        'batches': batches,
        'class_filter': class_filter,
        'batch_filter': batch_filter,
        'search_query': search_query,
        'form': form,
    }
    return render(request, 'college/fees.html', context)

@user_passes_test(is_admin, login_url='dashboard')
def add_fee_payment(request):
    if request.method == 'POST':
        form = FeePaymentForm(request.POST)
        if form.is_valid():
            payment = form.save()
            
            # Auto-generate AccountTransaction entry for this Income
            AccountTransaction.objects.create(
                transaction_type='INCOME',
                category='Course Fees',
                amount=payment.amount_paid,
                date=payment.payment_date,
                payment_method=payment.payment_method,
                description=f"Fee collected from Student: {payment.student.full_name} ({payment.student.roll_number}). Receipt: {payment.receipt_number}"
            )
            
            messages.success(request, f"Fee payment of {payment.amount_paid} logged successfully for {payment.student.full_name}.")
            return redirect('receipt_detail', pk=payment.pk)
    return redirect('fee_collection')

@user_passes_test(is_admin, login_url='dashboard')
def fee_defaulters(request):
    class_filter = request.GET.get('class_name', '')
    batch_filter = request.GET.get('batch', '')
    
    all_students = Student.objects.filter(is_active=True)
    if class_filter:
        all_students = all_students.filter(class_name__iexact=class_filter)
    if batch_filter:
        all_students = all_students.filter(batch__iexact=batch_filter)
        
    defaulters = []
    for s in all_students:
        if s.pending_fee > 0:
            defaulters.append(s)
            
    classes = Student.objects.values_list('class_name', flat=True).distinct()
    batches = Student.objects.values_list('batch', flat=True).distinct()
            
    context = {
        'students': defaulters,
        'classes': classes,
        'batches': batches,
        'class_filter': class_filter,
        'batch_filter': batch_filter,
    }
    return render(request, 'college/defaulters.html', context)

@user_passes_test(is_admin, login_url='dashboard')
def receipt_detail(request, pk):
    payment = get_object_or_404(FeeTransaction, pk=pk)
    student = payment.student
    
    # Generate WhatsApp prefilled URL
    msg = (
        f"*FEE RECEIPT - {student.class_name}*\n"
        f"College: College Management System\n"
        f"Student: {student.full_name}\n"
        f"Roll No: {student.roll_number}\n"
        f"Receipt: {payment.receipt_number}\n"
        f"Date: {payment.payment_date.strftime('%d-%m-%Y')}\n"
        f"Paid Amount: Rs. {payment.amount_paid}\n"
        f"Remaining Balance: Rs. {student.pending_fee}\n"
        f"Thank you!"
    )
    import urllib.parse
    encoded_msg = urllib.parse.quote(msg)
    whatsapp_url = f"https://api.whatsapp.com/send?phone={student.parent_phone}&text={encoded_msg}"
    
    return render(request, 'college/receipt_detail.html', {
        'payment': payment,
        'student': student,
        'whatsapp_url': whatsapp_url,
    })


# ----------------- DAILY ACCOUNTS -----------------
@user_passes_test(is_admin, login_url='dashboard')
def accounts_list(request):
    today = timezone.localtime(timezone.now()).date()
    
    # Filter
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    category_filter = request.GET.get('category', '')
    type_filter = request.GET.get('type', '')
    
    transactions = AccountTransaction.objects.all().order_by('-date')
    
    if start_date_str:
        transactions = transactions.filter(date__gte=start_date_str)
    if end_date_str:
        transactions = transactions.filter(date__lte=end_date_str)
    if category_filter:
        transactions = transactions.filter(category__icontains=category_filter)
    if type_filter:
        transactions = transactions.filter(transaction_type=type_filter)
        
    # Summary math
    total_income = transactions.filter(transaction_type='INCOME').aggregate(total=Sum('amount'))['total'] or 0
    total_expense = transactions.filter(transaction_type='EXPENSE').aggregate(total=Sum('amount'))['total'] or 0
    net_profit = total_income - total_expense
    
    if request.method == 'POST':
        form = AccountTransactionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Daily transaction logged successfully.")
            return redirect('accounts_list')
    else:
        form = AccountTransactionForm(initial={'date': today})
        
    categories = AccountTransaction.objects.values_list('category', flat=True).distinct()
    
    context = {
        'transactions': transactions,
        'form': form,
        'total_income': total_income,
        'total_expense': total_expense,
        'net_profit': net_profit,
        'categories': categories,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'category_filter': category_filter,
        'type_filter': type_filter,
    }
    return render(request, 'college/accounts.html', context)


# ----------------- SALARY MANAGEMENT -----------------
@user_passes_test(is_admin, login_url='dashboard')
def salary_list(request):
    today = timezone.localtime(timezone.now()).date()
    selected_month = request.GET.get('month', today.strftime('%Y-%m'))
    
    staff_members = User.objects.filter(profile__role='STAFF')
    payments = SalaryPayment.objects.filter(month=selected_month)
    payment_map = {pay.staff_id: pay for pay in payments}
    
    if request.method == 'POST':
        form = SalaryPaymentForm(request.POST)
        if form.is_valid():
            payment = form.save()
            
            # If paid, log in AccountTransaction automatically as EXPENSE
            if payment.payment_status == 'PAID':
                AccountTransaction.objects.create(
                    transaction_type='EXPENSE',
                    category='Staff Salary',
                    amount=payment.amount_paid,
                    date=payment.payment_date,
                    payment_method='BANK',  # Defaults to bank payment for salary
                    description=f"Salary paid to: {payment.staff.get_full_name() or payment.staff.username} for {payment.month}."
                )
            messages.success(request, f"Salary payment for {payment.staff.username} saved.")
            return redirect(f"{request.path}?month={selected_month}")
            
    staff_list = []
    for staff in staff_members:
        basic_sal = staff.profile.salary if hasattr(staff, 'profile') else 0.00
        payment = payment_map.get(staff.id, None)
        
        staff_list.append({
            'user': staff,
            'basic_salary': basic_sal,
            'payment': payment,
        })
        
    form = SalaryPaymentForm(initial={'payment_date': today, 'month': selected_month})
    
    context = {
        'staff_list': staff_list,
        'selected_month': selected_month,
        'form': form,
    }
    return render(request, 'college/salary.html', context)

@user_passes_test(is_admin, login_url='dashboard')
def pay_salary(request):
    if request.method == 'POST':
        staff_id = request.POST.get('staff')
        month = request.POST.get('month')
        basic_salary = float(request.POST.get('basic_salary', 0))
        advance_deducted = float(request.POST.get('advance_deducted', 0))
        other_deductions = float(request.POST.get('other_deductions', 0))
        amount_paid = basic_salary - advance_deducted - other_deductions
        payment_date = request.POST.get('payment_date', timezone.localtime(timezone.now()).date())
        remarks = request.POST.get('remarks', '')
        
        staff = get_object_or_404(User, id=staff_id)
        
        # Save payment statement
        payment, created = SalaryPayment.objects.update_or_create(
            staff=staff,
            month=month,
            defaults={
                'basic_salary': basic_salary,
                'advance_deducted': advance_deducted,
                'other_deductions': other_deductions,
                'amount_paid': amount_paid,
                'payment_date': payment_date,
                'payment_status': 'PAID',
                'remarks': remarks
            }
        )
        
        # Log as accounts expense
        AccountTransaction.objects.create(
            transaction_type='EXPENSE',
            category='Staff Salary',
            amount=amount_paid,
            date=payment_date,
            payment_method='BANK',
            description=f"Salary statement paid for {staff.get_full_name() or staff.username} ({month})"
        )
        messages.success(request, f"Salary statement of Rs.{amount_paid} recorded for {staff.username}.")
    return redirect(f"/salary/?month={month}")


# ----------------- REPORTS EXPORT (EXCEL) -----------------
@login_required
def export_attendance_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Headers
    headers = ["Student Name", "Roll Number", "Class", "Batch", "Date", "Period", "Status", "Marked By"]
    ws.append(headers)
    
    # Style
    font = Font(name="Arial", size=11, bold=True)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        
    # Filters
    class_filter = request.GET.get('class_name', '')
    batch_filter = request.GET.get('batch', '')
    date_str = request.GET.get('date', '')
    
    records = Attendance.objects.all()
    if class_filter:
        records = records.filter(student__class_name__iexact=class_filter)
    if batch_filter:
        records = records.filter(student__batch__iexact=batch_filter)
    if date_str:
        records = records.filter(date=date_str)
        
    for r in records:
        ws.append([
            r.student.full_name,
            r.student.roll_number,
            r.student.class_name,
            r.student.batch,
            r.date.strftime('%Y-%m-%d'),
            r.period,
            r.status,
            r.marked_by.username if r.marked_by else 'N/A'
        ])
        
    wb.save(response)
    return response

@login_required
def export_fees_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="fees_report.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fees Ledger"
    
    headers = ["Student Name", "Roll Number", "Class", "Batch", "Course Fee", "Total Paid", "Pending Fee", "Status"]
    ws.append(headers)
    
    # Style
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        
    class_filter = request.GET.get('class_name', '')
    batch_filter = request.GET.get('batch', '')
    
    students = Student.objects.filter(is_active=True)
    if class_filter:
        students = students.filter(class_name__iexact=class_filter)
    if batch_filter:
        students = students.filter(batch__iexact=batch_filter)
        
    for s in students:
        status_text = "PAID" if s.pending_fee <= 0 else f"PENDING (Rs. {s.pending_fee})"
        ws.append([
            s.full_name,
            s.roll_number,
            s.class_name,
            s.batch,
            float(s.total_course_fee),
            float(s.paid_fee),
            float(s.pending_fee),
            status_text
        ])
        
    wb.save(response)
    return response

@user_passes_test(is_admin, login_url='dashboard')
def export_accounts_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="accounts_report.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Transactions"
    
    headers = ["Date", "Type", "Category", "Amount", "Method", "Description"]
    ws.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    records = AccountTransaction.objects.all().order_by('date')
    if start_date:
        records = records.filter(date__gte=start_date)
    if end_date:
        records = records.filter(date__lte=end_date)
        
    for r in records:
        ws.append([
            r.date.strftime('%Y-%m-%d'),
            r.transaction_type,
            r.category,
            float(r.amount),
            r.payment_method,
            r.description or ''
        ])
        
    wb.save(response)
    return response


# ----------------- REPORTS EXPORT (PDF) -----------------
@login_required
def export_attendance_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=15
    )
    meta_style = ParagraphStyle(
        'MetaStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=20
    )
    
    story.append(Paragraph("Student Attendance Report", title_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
    
    class_filter = request.GET.get('class_name', '')
    batch_filter = request.GET.get('batch', '')
    date_str = request.GET.get('date', '')
    
    records = Attendance.objects.all().order_by('student__roll_number')
    if class_filter:
        records = records.filter(student__class_name__iexact=class_filter)
    if batch_filter:
        records = records.filter(student__batch__iexact=batch_filter)
    if date_str:
        records = records.filter(date=date_str)
        
    # Table data
    data = [["Roll No", "Student Name", "Class", "Batch", "Date", "Period", "Status"]]
    for r in records:
        data.append([
            r.student.roll_number,
            r.student.full_name,
            r.student.class_name,
            r.student.batch,
            r.date.strftime('%Y-%m-%d'),
            r.period,
            r.status
        ])
        
    t = Table(data, colWidths=[60, 140, 60, 60, 70, 70, 70])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('FONTSIZE', (0,1), (-1,-1), 9),
    ]))
    story.append(t)
    
    doc.build(story)
    return response

@login_required
def export_fees_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="fees_report.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#2E7D32'),
        spaceAfter=15
    )
    meta_style = ParagraphStyle(
        'MetaStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=20
    )
    
    story.append(Paragraph("Fee Collection & Outstanding Ledger", title_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
    
    class_filter = request.GET.get('class_name', '')
    batch_filter = request.GET.get('batch', '')
    
    students = Student.objects.filter(is_active=True).order_by('roll_number')
    if class_filter:
        students = students.filter(class_name__iexact=class_filter)
    if batch_filter:
        students = students.filter(batch__iexact=batch_filter)
        
    data = [["Roll No", "Student Name", "Class", "Batch", "Course Fee", "Total Paid", "Pending Fee"]]
    for s in students:
        data.append([
            s.roll_number,
            s.full_name,
            s.class_name,
            s.batch,
            f"Rs. {s.total_course_fee}",
            f"Rs. {s.paid_fee}",
            f"Rs. {s.pending_fee}"
        ])
        
    t = Table(data, colWidths=[60, 150, 60, 60, 70, 70, 70])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('FONTSIZE', (0,1), (-1,-1), 9),
    ]))
    story.append(t)
    doc.build(story)
    return response

@user_passes_test(is_admin, login_url='dashboard')
def export_accounts_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="financial_report.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#37474F'),
        spaceAfter=15
    )
    meta_style = ParagraphStyle(
        'MetaStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=20
    )
    
    story.append(Paragraph("Daily Accounts & Financial Transactions Report", title_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
    
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    records = AccountTransaction.objects.all().order_by('date')
    if start_date:
        records = records.filter(date__gte=start_date)
    if end_date:
        records = records.filter(date__lte=end_date)
        
    data = [["Date", "Type", "Category", "Amount", "Method", "Description"]]
    for r in records:
        desc = (r.description[:25] + '...') if (r.description and len(r.description) > 28) else (r.description or '')
        data.append([
            r.date.strftime('%Y-%m-%d'),
            r.transaction_type,
            r.category,
            f"Rs. {r.amount}",
            r.payment_method,
            desc
        ])
        
    t = Table(data, colWidths=[70, 60, 90, 70, 70, 180])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#37474F')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-2), 'CENTER'),  # Center all except description
        ('ALIGN', (5,0), (5,-1), 'LEFT'),   # Left align description
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('FONTSIZE', (0,1), (-1,-1), 9),
    ]))
    story.append(t)
    doc.build(story)
    return response
